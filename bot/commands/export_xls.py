import datetime
import os
import tempfile

from openpyxl import Workbook
from openpyxl.styles import Font, colors
from telegram import Update
from telegram.ext import CallbackContext

from bot import session, bot
from bot.conversations.statistics.utils import get_lifetime_user
from bot.models import User, Consumption, CategoryConsumption, Earning, CategoryEarning
from bot.utils import update_username, log_handler, update_activity


@update_username
@log_handler
@update_activity
def handler_export(update: Update, context: CallbackContext):
    user = session.query(User).filter(
        User.telegram_user_id == update.message.from_user.id
    ).first()
    xls_maker = XlsMaker(user)
    xls = xls_maker.make_xlx_file()
    with tempfile.TemporaryDirectory() as tmpdirname:
        save_path = os.path.join(tmpdirname, "{}.xlsx".format(get_lifetime_user(user, datetime.datetime.now())))
        xls.save(save_path)
        bot.send_document(chat_id=update.message.from_user.id,
                          document=open(save_path, 'rb'))


class XlsMaker:
    def __init__(self, user):
        self.user = user
        self.wb = Workbook()
        self.ws = self.wb.active

    def make_xlx_file(self):
        self._make_head()
        self._fill_consumptions()
        self._fill_earnings()
        self._expand_cells()
        return self.wb

    def _fill_consumptions(self):
        consumptions = session.query(Consumption, CategoryConsumption.category).filter(
            Consumption.user_id == self.user.id,
            CategoryConsumption.id == Consumption.category_id
        ).order_by(Consumption.time_creation.asc()).all()
        for index, item in enumerate(consumptions):
            consumption, category = item
            row_number = index + 3
            self.ws.cell(row=row_number, column=1, value=self._process_time_creation(consumption.time_creation))
            self.ws.cell(row=row_number, column=2, value=category)
            self.ws.cell(row=row_number, column=3, value=consumption.amount_money)

    def _expand_cells(self):
        ws = self.ws
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value

    def _fill_earnings(self):
        earnings = session.query(Earning, CategoryEarning.category).filter(
            Consumption.user_id == self.user.id,
            CategoryEarning.id == Earning.category_id
        ).order_by(Earning.time_creation.asc()).all()
        for index, item in enumerate(earnings):
            consumption, category = item
            row_number = index + 3
            self.ws.cell(row=row_number, column=5, value=self._process_time_creation(consumption.time_creation))
            self.ws.cell(row=row_number, column=6, value=category)
            self.ws.cell(row=row_number, column=7, value=consumption.amount_money)

    def _make_head(self):
        self.ws.title = get_lifetime_user(self.user, datetime.datetime.now())
        ft_red = Font(color=colors.RED)
        ft_green = Font(color=colors.GREEN)
        self.ws['A1'] = 'РАСХОДЫ'
        self.ws['A2'] = 'Дата'
        self.ws['B2'] = 'Категория'
        self.ws['C2'] = 'Кол-во денег'
        self.ws['E1'] = 'ДОХОДЫ'
        self.ws['E2'] = 'Дата'
        self.ws['F2'] = 'Категория'
        self.ws['G2'] = 'Кол-во денег'
        for col in self.ws.columns:
            col[0].font = Font(bold=True)
            col[1].font = Font(bold=True)
        self.ws['A1'].font = ft_red
        self.ws['E1'].font = ft_green

    @staticmethod
    def _process_time_creation(datetime_):
        return f"{datetime_.strftime('%d.%m.%Y')} {datetime_.strftime('%H:%M')}"

